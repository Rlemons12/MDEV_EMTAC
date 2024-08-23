from flask import Blueprint, render_template, request, redirect, url_for, flash
import os
import logging
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from config import DB_LOADSHEET_BOMS
from config_env import DatabaseConfig
from emtacdb_fts import Position, add_image_to_db

bill_of_materials_bp = Blueprint('bill_of_materials_bp', __name__, template_folder='templates')

# Setup logging
log_file_path = os.path.join(DB_LOADSHEET_BOMS, 'bom_loadsheet.log')
logging.basicConfig(
    filename=log_file_path,
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

db_config = DatabaseConfig()

def allowed_file(filename):
    allowed = '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx'}
    logging.debug(f"File {filename} allowed: {allowed}")
    return allowed
@bill_of_materials_bp.route('/bill_of_materials', methods=['GET', 'POST'])
def bill_of_materials():
    logging.debug("Accessed bill_of_materials route.")
    if request.method == 'POST':
        logging.debug("Received POST request.")

        # Retrieve the image path from the form
        image_path = request.form.get('image_path')
        logging.debug(f"Image path received: {image_path}")
        if not image_path:
            flash('Image path is required')
            logging.error("Image path is required but not provided.")
            return redirect(request.url)

        # Retrieve data from the form for Position creation
        area_id = request.form.get('area')
        equipment_group_id = request.form.get('equipment_group')
        model_id = request.form.get('model')
        asset_number_id = request.form.get('asset_number')
        location_id = request.form.get('location')
        site_location_id = request.form.get('site_location')
        logging.debug(f"Received Position data: Area ID={area_id}, Equipment Group ID={equipment_group_id}, "
                      f"Model ID={model_id}, Asset Number ID={asset_number_id}, Location ID={location_id}, "
                      f"Site Location ID={site_location_id}")

        session = db_config.get_main_session()

        # Create a new Position object
        new_position = Position(
            area_id=area_id,
            equipment_group_id=equipment_group_id,
            model_id=model_id,
            asset_number_id=asset_number_id,
            location_id=location_id,
            site_location_id=site_location_id
        )

        try:
            session.add(new_position)
            session.commit()
            position_id = new_position.id  # Get the ID of the newly created Position
            flash('Position created successfully!')
            logging.info(f"Position created successfully with ID: {position_id}")
        except Exception as e:
            session.rollback()
            logging.error(f"Error creating position: {e}")
            flash(f"Error creating position: {str(e)}")
            return redirect(request.url)
        finally:
            session.close()

        # Check if the POST request has the file part
        if 'file' not in request.files:
            flash('No file part')
            logging.error("No file part in the POST request.")
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No selected file')
            logging.error("No selected file in the POST request.")
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(DB_LOADSHEET_BOMS, filename)
            file.save(file_path)
            flash('File successfully uploaded')
            logging.info(f"File successfully uploaded: {file_path}")

            try:
                process_bom_loadsheet(file_path, image_path, position_id)  # Pass the position_id here
                flash('File successfully processed')
                logging.info(f"File successfully processed: {file_path}")
            except Exception as e:
                logging.error(f"Error processing file {file_path}: {e}")
                flash(f"An error occurred while processing the file: {e}")

        return redirect(url_for('bill_of_materials_bp.bill_of_materials'))

    return render_template('bill_of_materials.html')

def process_bom_loadsheet(source_file, image_path, position_id):
    logging.info(f"Starting BOM loadsheet process for file: {source_file}")

    # Generate the target file path based on the uploaded file
    target_path = prompt_for_target_file(source_file)

    # Copy the BOM sheet to the new sheet in the target workbook
    copy_bom_sheet_to_target(source_file, target_path)

    # Match items and update the part_position_image sheet
    match_items_and_update_sheet(target_path, image_path, position_id)

    logging.info("BOM loadsheet process completed.")

def prompt_for_target_file(source_file):
    logging.info(f"Generating target file name based on source file: {source_file}")
    suffix = os.path.basename(source_file).replace("bom_for_", "").replace(".xlsx", "")

    if not os.path.exists(DB_LOADSHEET_BOMS):
        os.makedirs(DB_LOADSHEET_BOMS)
        logging.info(f"Directory {DB_LOADSHEET_BOMS} created.")

    target_file_name = f"load_bom_for_{suffix}.xlsx"
    file_path = os.path.join(DB_LOADSHEET_BOMS, target_file_name)

    if os.path.exists(file_path):
        logging.info(f"Target file already exists: {file_path}")
        print(f"The file {file_path} already exists and will be updated.")
    else:
        logging.info(f"Creating new target file: {file_path}")
        print(f"The file {file_path} does not exist. A new workbook will be created.")

    return file_path


def copy_bom_sheet_to_target(source_path, target_path):
    logging.info(f"Copying BOM sheet from {source_path} to {target_path}.")
    wb_source = load_workbook(source_path)
    if "BOM" in wb_source.sheetnames:
        bom_sheet = wb_source["BOM"]

        if os.path.exists(target_path):
            wb_target = load_workbook(target_path)
        else:
            wb_target = Workbook()
            default_sheet = wb_target.active
            wb_target.remove(default_sheet)

        suffix = os.path.basename(target_path).replace("load_bom_for_", "").replace(".xlsx", "")
        new_bom_sheet_name = f"bom_{suffix}"
        bom_target_sheet = wb_target.create_sheet(new_bom_sheet_name)

        for row in bom_sheet.iter_rows(values_only=True):
            bom_target_sheet.append(row)

        part_position_image_sheet = wb_target.create_sheet("part_position_image")
        part_position_image_sheet.append(["part", "position", "image", "description"])

        wb_target.save(target_path)
        logging.info(f"'BOM' sheet copied to '{new_bom_sheet_name}' in {target_path}.")
    else:
        logging.error(f"The source workbook does not contain a sheet named 'BOM'. Source: {source_path}")
        print("The source workbook does not contain a sheet named 'BOM'.")

def process_row(part_position_image_sheet, item_number, photo, description, manufacturer_description, position_id):
    logging.debug(f"Processing row: Item Number: {item_number}, Photo: {photo}")
    if photo:
        full_description = f"{description}, {manufacturer_description}" if manufacturer_description else description
        prefixed_photo = f"A{photo}"  # Prefix the photo with "A" inside process_row
        part_position_image_sheet.append([item_number, position_id, prefixed_photo, full_description])
        logging.info(
            f"Added entry: Item Number: {item_number}, Position ID: {position_id}, Photo: {prefixed_photo}, Description: {full_description}")


def match_items_and_update_sheet(target_path, image_path, position_id):
    logging.info(f"Matching items and updating part_position_image sheet in {target_path}.")
    wb_target = load_workbook(target_path)
    bom_sheet_name = [sheet for sheet in wb_target.sheetnames if sheet.startswith("bom_")][0]
    bom_sheet = wb_target[bom_sheet_name]
    part_position_image_sheet = wb_target["part_position_image"]

    part_list_image_path = os.path.join(DB_LOADSHEET_BOMS, "part_list_image.xlsx")
    wb_part_list = load_workbook(part_list_image_path)
    photo_list_sheet = wb_part_list["photo_list"]

    match_count = 0  # Counter to limit processing to the first 5 matches

    for row in bom_sheet.iter_rows(min_row=2, values_only=True):
        item_number = str(row[3])  # Assuming "Item Number" is the fourth column in BOM sheet

        if item_number and item_number.startswith("A"):
            item_number = item_number[1:]

        part_number_prefix = item_number[:6]

        for photo_row in photo_list_sheet.iter_rows(min_row=2, values_only=True):
            photo_part_number_prefix = str(photo_row[0])[:6]  # First 6 characters of Part #
            if part_number_prefix == photo_part_number_prefix:
                match_count += 1

                photo_a = photo_row[1]
                desc_a = photo_row[4]  # Corresponding "Desc A"
                photo_b = photo_row[2]
                desc_b = photo_row[5]  # Corresponding "Desc B"
                photo_c = photo_row[3]
                desc_c = photo_row[6]  # Corresponding "Desc C"
                manufacturer_description = photo_row[7]  # "Manufacturer Description"

                process_row(part_position_image_sheet, item_number, photo_a, desc_a, manufacturer_description, position_id)
                process_row(part_position_image_sheet, item_number, photo_b, desc_b, manufacturer_description, position_id)
                process_row(part_position_image_sheet, item_number, photo_c, desc_c, manufacturer_description, position_id)

                # Add logging and pass the original photo names to process_part_position_image with the "A" prefix
                prefixed_photo_a = f"A{photo_a}"
                prefixed_photo_b = f"A{photo_b}"
                prefixed_photo_c = f"A{photo_c}"

                logging.debug(f"Passing prefixed photo name to process_part_position_image: {prefixed_photo_a}")
                process_part_position_image(item_number, position_id, prefixed_photo_a, image_path)

                logging.debug(f"Passing prefixed photo name to process_part_position_image: {prefixed_photo_b}")
                process_part_position_image(item_number, position_id, prefixed_photo_b, image_path)

                logging.debug(f"Passing prefixed photo name to process_part_position_image: {prefixed_photo_c}")
                process_part_position_image(item_number, position_id, prefixed_photo_c, image_path)

                if match_count >= 5:
                    logging.info("Reached the limit of 5 matches. Stopping further processing.")
                    break

        if match_count >= 5:
            break

    wb_target.save(target_path)
    logging.info(f"part_position_image sheet updated in {target_path}.")

def find_image_in_subfolders(image_title, base_path):
    logging.debug(f"Searching for image '{image_title}' in '{base_path}' and its subfolders.")
    """
    Search for the image in the given base_path and its subfolders.
    Returns the full file path if found, otherwise returns None.
    """
    for root, dirs, files in os.walk(base_path):
        logging.debug(f"Searching in directory: {root}")  # Log each directory being searched
        for file in files:
            # Get the file name without extension
            file_name_without_ext = os.path.splitext(file)[0]
            # Check if the file name matches the image_title
            if file_name_without_ext == image_title:
                full_path = os.path.join(root, file)
                logging.info(f"Image found: {full_path}")
                return full_path
    logging.warning(f"Image '{image_title}' not found in '{base_path}' or its subfolders.")
    return None

def process_part_position_image(part_number, position_value, image_title, base_image_path):
    logging.info(
        f"Starting to process part position image for Part Number: {part_number}, Position Value: {position_value}, Image Title: {image_title}"
    )

    # Locate the image file in the base path or its subfolders
    logging.debug(f"Attempting to locate image '{image_title}' in '{base_image_path}' or its subfolders.")
    image_file_path = find_image_in_subfolders(image_title, base_image_path)

    if not image_file_path:
        logging.error(f"Image '{image_title}' not found in {base_image_path} or its subfolders. Aborting process for this image.")
        return

    # Add the image to the database if it doesn't exist
    logging.debug(f"Found image '{image_title}' at '{image_file_path}'. Attempting to add to the database.")
    image_id = add_image_to_db(title=image_title, file_path=image_file_path, position_id=position_value)

    if image_id:
        logging.info(f"Image '{image_title}' processed and added to the database with ID {image_id}.")
    else:
        logging.error(f"Failed to process image '{image_title}'. It was not added to the database.")

    logging.info(f"Finished processing part position image for Part Number: {part_number}, Position Value: {position_value}, Image Title: {image_title}")
