import os
import sys
import logging
from openpyxl import load_workbook, Workbook
from concurrent.futures import ThreadPoolExecutor
from modules.configuration.config import DB_LOADSHEET_BOMS

# Append parent directory to system path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# Setup logging
log_file_path = os.path.join(DB_LOADSHEET_BOMS, 'bom_loadsheet.log')
logging.basicConfig(
    filename=log_file_path,
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def prompt_for_source_file():
    logging.info("Prompting for source file.")
    file_path = input("Please enter the path to the source file that starts with 'bom_for_': ")
    file_path = file_path.strip('\"\'')  # Remove extra quotes

    if os.path.exists(file_path) and os.path.basename(file_path).startswith("bom_for_"):
        logging.info(f"Source file found: {file_path}")
        return file_path
    else:
        logging.error(f"File does not exist or does not start with 'bom_for_': {file_path}")
        print("The file does not exist or does not start with 'bom_for_'. Please try again.")
        return prompt_for_source_file()

def prompt_for_target_file(source_file):
    logging.info("Generating target file name based on source file.")
    # Extract the suffix from the source file name
    suffix = os.path.basename(source_file).replace("bom_for_", "").replace(".xlsx", "")

    # Ensure target directory exists
    if not os.path.exists(DB_LOADSHEET_BOMS):
        os.makedirs(DB_LOADSHEET_BOMS)
        logging.info(f"Directory {DB_LOADSHEET_BOMS} created.")

    # Generate target file path
    target_file_name = f"load_bom_for_{suffix}.xlsx"
    file_path = os.path.join(DB_LOADSHEET_BOMS, target_file_name)

    # Log whether the file already exists
    if os.path.exists(file_path):
        logging.info(f"Target file already exists: {file_path}")
        print(f"The file {file_path} already exists and will be updated.")
    else:
        logging.info(f"Creating new target file: {file_path}")
        print(f"The file {file_path} does not exist. A new workbook will be created.")

    return file_path

def copy_bom_sheet_to_target(source_path, target_path):
    logging.info(f"Copying BOM sheet from {source_path} to {target_path}.")
    wb_source = load_workbook(source_path)

    # Check if the BOM sheet exists in the source file
    if "BOM" in wb_source.sheetnames:
        bom_sheet = wb_source["BOM"]

        # Open or create the target workbook
        if os.path.exists(target_path):
            wb_target = load_workbook(target_path)
            logging.info(f"Opened existing target workbook: {target_path}")
        else:
            wb_target = Workbook()
            default_sheet = wb_target.active
            wb_target.remove(default_sheet)
            logging.info(f"Created new target workbook: {target_path}")

        # Create the new BOM sheet in the target workbook
        suffix = os.path.basename(target_path).replace("load_bom_for_", "").replace(".xlsx", "")
        new_bom_sheet_name = f"bom_{suffix}"
        bom_target_sheet = wb_target.create_sheet(new_bom_sheet_name)
        logging.info(f"Created new sheet: {new_bom_sheet_name}")

        # Copy the content of the BOM sheet to the new sheet
        for row in bom_sheet.iter_rows(values_only=True):
            bom_target_sheet.append(row)
        logging.info(f"Copied BOM sheet data to {new_bom_sheet_name} in target workbook.")

        # Create the part_position_image sheet if it doesn't exist
        if "part_position_image" not in wb_target.sheetnames:
            part_position_image_sheet = wb_target.create_sheet("part_position_image")
            part_position_image_sheet.append(["part", "position", "image", "description"])
            logging.info(f"Created new sheet: part_position_image")

        # Save the target workbook
        wb_target.save(target_path)
        logging.info(f"Saved target workbook: {target_path}")
    else:
        logging.error(f"The source workbook does not contain a sheet named 'BOM'.")
        print("The source workbook does not contain a sheet named 'BOM'.")

def process_row(part_position_image_sheet, item_number, photo, description, manufacturer_description):
    if photo:
        full_description = f"{description}, {manufacturer_description}" if manufacturer_description else description
        part_position_image_sheet.append([item_number, "", photo, full_description])
        logging.info(f"Added entry: Item Number: {item_number}, Photo: {photo}, Description: {full_description}")

def match_items_and_update_sheet(target_path):
    logging.info("Matching items and updating part_position_image sheet.")
    wb_target = load_workbook(target_path)
    bom_sheet_name = [sheet for sheet in wb_target.sheetnames if sheet.startswith("bom_")][0]
    bom_sheet = wb_target[bom_sheet_name]
    part_position_image_sheet = wb_target["part_position_image"]

    part_list_image_path = os.path.join(DB_LOADSHEET_BOMS, "part_list_image.xlsx")
    wb_part_list = load_workbook(part_list_image_path)
    photo_list_sheet = wb_part_list["photo_list"]

    # Logging headers of the BOM and photo list sheets
    bom_headers = [cell.value for cell in bom_sheet[1]]
    photo_list_headers = [cell.value for cell in photo_list_sheet[1]]
    logging.debug(f"BOM sheet headers: {bom_headers}")
    logging.debug(f"Photo list sheet headers: {photo_list_headers}")

    match_count = 0  # Counter to limit processing to the first 5 matches

    with ThreadPoolExecutor() as executor:
        for row in bom_sheet.iter_rows(min_row=2, values_only=True):
            item_number = str(row[3])  # Assuming "Item Number" is the fourth column in bom sheet
            logging.debug(f"Processing Item Number: {item_number} (Full row data: {row})")

            if item_number and item_number.startswith("A"):
                item_number = item_number[1:]  # Remove leading 'A' if present
                logging.debug(f"Stripped leading 'A': {item_number}")

            part_number_prefix = item_number[:6]

            for photo_row in photo_list_sheet.iter_rows(min_row=2, values_only=True):
                photo_part_number_prefix = str(photo_row[0])[:6]  # First 6 characters of Part #
                if part_number_prefix == photo_part_number_prefix:
                    logging.info(f"Match found: Item Number: {item_number} - Part Number: {photo_row[0]}")
                    match_count += 1

                    # Retrieve the photo columns and corresponding descriptions
                    photo_a = photo_row[1]
                    desc_a = photo_row[4]  # Corresponding "Desc A"
                    photo_b = photo_row[2]
                    desc_b = photo_row[5]  # Corresponding "Desc B"
                    photo_c = photo_row[3]
                    desc_c = photo_row[6]  # Corresponding "Desc C"
                    manufacturer_description = photo_row[7]  # "Manufacturer Description"

                    # Process each row in a separate thread
                    executor.submit(process_row, part_position_image_sheet, item_number, photo_a, desc_a, manufacturer_description)
                    executor.submit(process_row, part_position_image_sheet, item_number, photo_b, desc_b, manufacturer_description)
                    executor.submit(process_row, part_position_image_sheet, item_number, photo_c, desc_c, manufacturer_description)

                    if match_count >= 5:
                        logging.info("Reached the limit of 5 matches. Stopping further processing.")
                        break

            if match_count >= 5:
                break

    wb_target.save(target_path)
    logging.info(f"part_position_image sheet updated in {target_path}.")

def main():
    logging.info("Starting bom_loadsheet process.")
    source_path = prompt_for_source_file()
    target_path = prompt_for_target_file(source_path)
    copy_bom_sheet_to_target(source_path, target_path)
    match_items_and_update_sheet(target_path)
    logging.info("bom_loadsheet process completed.")

if __name__ == "__main__":
    main()
